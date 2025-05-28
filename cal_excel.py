import openpyxl
from glob import glob

wb = openpyxl.load_workbook("excel/bill_list.xlsx")
ws = wb["Sheet1"]

wb_test1 = openpyxl.load_workbook("excel/bill_test1.xlsx", data_only=True)
ws_test1 = wb_test1["Sheet1"]

files = sorted(glob("excel/bill_test*.xlsx"))

for file in files:

    wb_test = openpyxl.load_workbook(file, data_only=True)
    ws_test = wb_test.worksheets[0]

    line = ws.max_row + 1
    ws.cell(line, 1).value = ws_test.cell(3, 14).value
    ws.cell(line, 2).value = ws_test.cell(3, 1).value
    ws.cell(line, 3).number_format = r"¥#,##0;¥-#,##0"
    ws.cell(line, 3).value = ws_test.cell(15, 4).value
    ws.cell(line, 4).number_format = r"yyyy/m/d"
    ws.cell(line, 4).value = ws_test.cell(4, 14).value

# wb_test2 = openpyxl.load_workbook("excel/bill_test2.xlsx", data_only=True)
# ws_test2 = wb_test2.worksheets[0]

# wb_test3 = openpyxl.load_workbook("excel/bill_test3.xlsx")
# ws_test3 = wb_test3.worksheets[0]

# ws.cell(2, 1).value = ws_test1.cell(3, 14).value
# ws.cell(2, 2).value = ws_test1.cell(3, 1).value
# ws.cell(2, 3).number_format = r"¥#,##0;¥-#,##0"
# ws.cell(2, 3).value = ws_test1.cell(15, 4).value
# ws.cell(2, 4).number_format = r"yyyy/m/d"
# ws.cell(2, 4).value = ws_test1.cell(4, 14).value

# ws.cell(3, 1).value = ws_test2.cell(3, 14).value
# ws.cell(3, 2).value = ws_test2.cell(3, 1).value
# ws.cell(3, 3).number_format = r"¥#,##0;¥-#,##0"
# ws.cell(3, 3).value = ws_test2.cell(15, 4).value
# ws.cell(3, 4).number_format = r"yyyy/m/d"
# ws.cell(3, 4).value = ws_test2.cell(4, 14).value

wb.save("excel/bill_list.xlsx")